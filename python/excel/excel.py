

# openpyxl
import openpyxl

# Workbookクラス: ワークブック全体
# Worksheetクラス: 一つのシート
# Cellクラス: 一つのセル

wb = openpyxl.load_workbook('test.xlsx')

	# シートの中身
	# A	B	C	D
	# 123	てすと１	＠	＠
	# 12.3	てすと２	＠	＠
	# 1.23	てすと３	＠	＠

print(wb) # <openpyxl.workbook.workbook.Workbook object at 0x00000225A6E260A0>

print(wb.sheetnames) # ['Sheet1', 'Sheet2']

ws = wb['Sheet1']
print(ws) # <Worksheet "Sheet1">

cell = ws['A2']
print(cell) # <Cell 'Sheet1'.A2>
print(cell.value) # 123

# 行番号、列番号指定（1始まり）
cell = ws.cell(row=2, column=1)
print(cell.value) # 123

g = ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=3)
print(list(g))
# [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>),
#  (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>),
#  (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>),
#  (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>)]

# ワークシートの追加
ws_new = wb.create_sheet('sheet_new')
print(ws_new) # <Worksheet "sheet_new">
print(wb.worksheets) # [<Worksheet "Sheet1">, <Worksheet "Sheet2">, <Worksheet "sheet_new">]